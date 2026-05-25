import os
import re
import json
from typing import Optional

from time import sleep
from dotenv import load_dotenv
from openpyxl import load_workbook

from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage

load_dotenv()

api_key = os.getenv("GROQ_API_KEY")
chat_model = os.getenv("CHAT_MODEL")

raw_llm = ChatGroq(
    api_key=api_key,
    model=chat_model,
    temperature=0
)

# -----------------------------------
# CONFIG
# -----------------------------------

PM2_LOG_PATH = "/home/amartya-mandal/.pm2/logs/gps-server-out.log"
IMEI_XLSX_PATH = "imei_list.xlsx"

STATE_FILE = "imei_state.json"
LOG_POSITION_FILE = "log_position.txt"

# -----------------------------------
# CHAT MODEL
# -----------------------------------

def chat_with_model(user_message: str) -> str:

    response = raw_llm.invoke(
        [HumanMessage(content=user_message)]
    )

    return response.content

# -----------------------------------
# LOAD IMEI LIST
# -----------------------------------

def load_imei_list(path: str = IMEI_XLSX_PATH):

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"IMEI Excel file not found: {path}"
        )

    wb = load_workbook(path)

    ws = wb.active

    imei_list = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:
            imei_list.append(
                str(row[0]).strip()
            )

    return imei_list

# -----------------------------------
# LOAD STATE
# -----------------------------------

def load_state():

    if not os.path.exists(STATE_FILE):
        return {}

    with open(STATE_FILE, "r") as f:
        return json.load(f)

# -----------------------------------
# SAVE STATE
# -----------------------------------

def save_state(state):

    with open(STATE_FILE, "w") as f:
        json.dump(
            state,
            f,
            indent=2
        )

# -----------------------------------
# READ ONLY NEW LOGS
# -----------------------------------

def read_new_logs():

    last_position = 0

    if os.path.exists(LOG_POSITION_FILE):

        with open(LOG_POSITION_FILE, "r") as f:

            pos = f.read().strip()

            if pos:
                last_position = int(pos)

    with open(
        PM2_LOG_PATH,
        "r",
        encoding="utf-8",
        errors="replace"
    ) as f:

        f.seek(last_position)

        new_logs = f.read()

        current_position = f.tell()

    with open(LOG_POSITION_FILE, "w") as f:
        f.write(str(current_position))

    return new_logs

# -----------------------------------
# DIRECTION MAPPING
# -----------------------------------

def get_direction(course: int) -> str:

    try:
        course = int(course)
    except:
        return "Unknown"

    course = course % 360

    if 0 <= course <= 22:
        return "North"

    elif 23 <= course <= 67:
        return "North-East"

    elif 68 <= course <= 112:
        return "East"

    elif 113 <= course <= 157:
        return "South-East"

    elif 158 <= course <= 202:
        return "South"

    elif 203 <= course <= 247:
        return "South-West"

    elif 248 <= course <= 292:
        return "West"

    elif 293 <= course <= 337:
        return "North-West"

    elif 338 <= course <= 360:
        return "North"

    return "Unknown"

# -----------------------------------
# EXTRACT GPS DATA
# -----------------------------------

def extract_latest_imei_locations(log_text: str):

    imei_list = load_imei_list()

    state = load_state()

    lines = log_text.splitlines()

    for line in lines:

        if "Location - IMEI:" not in line:
            continue

        for imei in imei_list:

            if imei not in line:
                continue

            timestamp_match = re.search(
                r'(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})',
                line
            )

            lat_match = re.search(
                r'Lat:\s*([0-9.]+)',
                line
            )

            lon_match = re.search(
                r'Lon:\s*([0-9.]+)',
                line
            )

            speed_match = re.search(
                r'Speed:\s*([0-9]+)',
                line
            )

            course_match = re.search(
                r'Course:\s*([0-9]+)',
                line
            )

            if not lat_match or not lon_match:
                continue

            timestamp = (
                timestamp_match.group(1)
                if timestamp_match else "Unknown"
            )

            lat = round(
                float(lat_match.group(1)),
                6
            )

            lon = round(
                float(lon_match.group(1)),
                6
            )

            speed = (
                int(speed_match.group(1))
                if speed_match else 0
            )

            course = (
                int(course_match.group(1))
                if course_match else 0
            )

            state[imei] = {
                "timestamp": timestamp,
                "direction": get_direction(course),
                "lat": lat,
                "lon": lon,
                "speed": speed,
                "course": course
            }

    save_state(state)

    return state

# -----------------------------------
# PRINT TABLE
# -----------------------------------

def print_imei_table(data: dict):

    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+"
    )

    print(
        "| IMEI            | Timestamp           | Direction   | Latitude   | Longitude  | Speed | Course |"
    )

    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+"
    )

    for imei, info in data.items():

        print(
            f"| {imei:<15} "
            f"| {info['timestamp']:<19} "
            f"| {info['direction']:<11} "
            f"| {info['lat']:<10.6f} "
            f"| {info['lon']:<10.6f} "
            f"| {info['speed']:<5} "
            f"| {info['course']:<6} |"
        )

    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+"
    )

# -----------------------------------
# MAIN
# -----------------------------------

if __name__ == "__main__":

    print("Simple AI Assistant")

    print(
        "Commands: 'show log', 'analyze log', 'quit'\n"
    )

    while True:

        user_input = input("You: ").strip()

        if not user_input:
            continue

        text = user_input.lower()

        # -------------------------------
        # EXIT
        # -------------------------------

        if any(
            word in text
            for word in [
                "quit",
                "exit",
                "bye",
                "tata",
                "q"
            ]
        ):

            print("Goodbye!")

            break

        # -------------------------------
        # SHOW RAW LOGS
        # -------------------------------

        if "show" in text:

            try:

                print(
                    "Assistant: Displaying new logs..."
                )

                sleep(2)

                print(read_new_logs())

            except Exception as e:

                print(f"Assistant: {e}")

            continue

        # -------------------------------
        # ANALYZE GPS LOGS
        # -------------------------------

        if any(
            word in text
            for word in [
                "analyze",
                "summarize",
                "extract",
                "pattern",
                "decode"
            ]
        ):

            try:

                print(
                    "Assistant: Analyzing log, please wait..."
                )

                sleep(2)

                log_text = read_new_logs()

                latest_locations = (
                    extract_latest_imei_locations(
                        log_text
                    )
                )

                print_imei_table(
                    latest_locations
                )

            except Exception as e:

                print(f"Assistant: {e}")

            print()

            continue

        # -------------------------------
        # DEFAULT CHAT
        # -------------------------------

        print(
            "Assistant:",
            chat_with_model(user_input)
        )

        print()