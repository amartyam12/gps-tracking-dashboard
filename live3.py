import os
import re
import json
from time import sleep
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage

load_dotenv()
api_key = os.getenv("GROQ_API_KEY")
chat_model = os.getenv("CHAT_MODEL")

raw_llm = ChatGroq(api_key=api_key, model=chat_model, temperature=0)

PM2_LOG_PATH = "/home/amartya-mandal/.pm2/logs/gps-server-out.log"
IMEI_XLSX_PATH = "imei_list.xlsx"
STATE_FILE = "imei_state.json"
LOG_POSITION_FILE = "log_position.txt"
REFRESH_INTERVAL = 1
ALERT_SPEED = 100

def chat_with_model(user_message: str) -> str:
    response = raw_llm.invoke([HumanMessage(content=user_message)])
    return response.content

def load_imei_list():
    wb = load_workbook(IMEI_XLSX_PATH)
    ws = wb.active
    imeis = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            imeis.append(str(row[0]).strip())
    return imeis

def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, "r") as f:
        return json.load(f)
def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)

def read_new_logs():
    last_position = 0
    if os.path.exists(LOG_POSITION_FILE):
        with open(LOG_POSITION_FILE, "r") as f:
            pos = f.read().strip()
            if pos:
                last_position = int(pos)
    with open(PM2_LOG_PATH, "r", encoding="utf-8", errors="replace") as f:
        f.seek(last_position)
        new_logs = f.read()
        current_position = f.tell()
    with open(LOG_POSITION_FILE, "w") as f:
        f.write(str(current_position))
    return new_logs

def get_direction(course):
    try:
        course = int(course)
    except:
        return "Unknown"
    course = course % 360
    directions = [
        (22, "North"),
        (67, "North-East"),
        (112, "East"),
        (157, "South-East"),
        (202, "South"),
        (247, "South-West"),
        (292, "West"),
        (337, "North-West"),
        (360, "North"),
    ]
    for limit, direction in directions:
        if course <= limit:
            return direction
    return "Unknown"

def detect_anomalies(imei, data):
    alerts = []
    # HIGH SPEED
    if data["speed"] >= ALERT_SPEED:
        alerts.append(f"⚠️ HIGH SPEED ({data['speed']})")
    # OFFLINE DETECTION (1 HOUR)
    try:
        timestamp = datetime.strptime(data["timestamp"], "%Y/%m/%d %H:%M:%S")
        seconds = (datetime.now() - timestamp).total_seconds()
        if seconds > 3600:
            alerts.append("🔴 OFFLINE")
    except:
        pass
    return alerts

def extract_latest_imei_locations(log_text):
    imei_list = load_imei_list()
    state = load_state()
    lines = log_text.splitlines()
    for line in lines:
        if "Location - IMEI:" not in line:
            continue
        for imei in imei_list:
            if imei not in line:
                continue
            timestamp_match = re.search(r"(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})", line)
            lat_match = re.search(r"Lat:\s*([0-9.]+)", line)
            lon_match = re.search(r"Lon:\s*([0-9.]+)", line)
            speed_match = re.search(r"Speed:\s*([0-9]+)", line)
            course_match = re.search(r"Course:\s*([0-9]+)", line)
            if not lat_match or not lon_match:
                continue
            timestamp = timestamp_match.group(1) if timestamp_match else "Unknown"
            lat = round(float(lat_match.group(1)), 6)
            lon = round(float(lon_match.group(1)), 6)
            speed = int(speed_match.group(1)) if speed_match else 0
            course = int(course_match.group(1)) if course_match else 0
            direction = get_direction(course)
            state[imei] = {
                "timestamp": timestamp,
                "direction": direction,
                "lat": lat,
                "lon": lon,
                "speed": speed,
                "course": course,
            }
    save_state(state)
    return state

def print_table(state):
    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+----------------------+"
    )
    print(
        "| IMEI            | Timestamp           | Direction   | Latitude   | Longitude  | Speed | Course | Alerts               |"
    )
    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+----------------------+"
    )
    for imei, data in state.items():
        alerts = detect_anomalies(imei, data)
        alert_text = ", ".join(alerts) if alerts else "OK"
        print(
            f"| {imei:<15} "
            f"| {data['timestamp']:<19} "
            f"| {data['direction']:<11} "
            f"| {data['lat']:<10.6f} "
            f"| {data['lon']:<10.6f} "
            f"| {data['speed']:<5} "
            f"| {data['course']:<6} "
            f"| {alert_text:<20} |"
        )
    print(
        "+-----------------+---------------------+-------------+------------+------------+-------+--------+----------------------+"
    )

def ai_summary_agent(state):
    try:
        summary_prompt = f"""
        Analyze this GPS fleet data.
        Detect:
        - high speed vehicles
        - inactive vehicles
        - suspicious patterns
        - movement insights
        Data:
        {json.dumps(state, indent=2)}
        Keep response short.
        """
        result = chat_with_model(summary_prompt)
        print("\n🤖 AI FLEET ANALYSIS\n")
        print(result)
    except Exception as e:
        print(f"AI Summary Error: {e}")

def live_dashboard():
    print("\nStarting LIVE GPS dashboard...")
    print("Press CTRL+C to stop.\n")
    while True:
        try:
            log_text = read_new_logs()
            if log_text.strip():
                extract_latest_imei_locations(log_text)
            state = load_state()
            os.system("clear")
            print("🚀 LIVE AGENTIC AI GPS DASHBOARD\n")
            print_table(state)
            print(f"\nRefreshing every {REFRESH_INTERVAL} seconds...")
            sleep(REFRESH_INTERVAL)
        except KeyboardInterrupt:
            print("\nDashboard stopped.")
            break
        except Exception as e:
            print(f"Dashboard Error: {e}")
            sleep(2)
if __name__ == "__main__":
    print("AGENTIC AI GPS MONITORING SYSTEM")

    while True:
        user_input = input("\nYou:").strip()
        text = user_input.lower()

        if any (word in text for word in  ["quit", "exit", "bye", "q"]):
            print("Goodbye!")
            break
        if any (word in text for word in ["live", "show", "current", "status"]):
            live_dashboard()
            continue
        if any (word in text for word in ["decode","analyze","summarize"]):
            logs = read_new_logs()
            extract_latest_imei_locations(logs)
            state = load_state()
            print_table(state)
            continue
        print("Jarvis:", chat_with_model(user_input))
